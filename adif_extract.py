"""
adif_extract.py
===============
Lightweight ADIF spreadsheet editor.  Extract QSOs from any ADIF file
(QRZ, LoTW, N3FJP, WSJT-X, …) to a formatted Excel workbook or CSV,
edit values directly in Excel, then convert back to ADIF.

Typical POTA / portable workflow
---------------------------------
  1.  Extract a date range to Excel:
          python adif_extract.py --adif wt8p.adi --date 2026-03-28 --output-xlsx pota.xlsx

  2.  Open pota.xlsx in Excel.  Correct values in any column.
      Delete columns you don't need (they will be absent from the ADIF).
      Delete rows you don't want to reimport.  Save.

  3.  Convert the edited Excel back to ADIF:
          python adif_extract.py --from-xlsx pota.xlsx --output-adif pota_corrected.adi

  4.  Import pota_corrected.adi into your logging program, or use as the
      --adif source for resolve_qrz_discrepancies.py.

Usage - extract mode (requires --adif)
---------------------------------------
    python adif_extract.py --adif qrz_export.adi
    python adif_extract.py --adif qrz_export.adi --date 2026-03-28
    python adif_extract.py --adif qrz_export.adi --after 2026-03-01 --before 2026-03-31
    python adif_extract.py --adif qrz_export.adi --preset lotw --output-xlsx lotw_check.xlsx
    python adif_extract.py --adif qrz_export.adi --fields GRIDSQUARE,STATE,CNTY
    python adif_extract.py --adif qrz_export.adi --date 2026-03-28 --output-xlsx pota.xlsx --no-csv

Usage - round-trip mode (requires --from-xlsx)
-----------------------------------------------
    python adif_extract.py --from-xlsx pota.xlsx --output-adif pota_fixed.adi

Date formats accepted by --after / --before / --date
------------------------------------------------------
    YYYY-MM-DD   (ISO, e.g. 2026-03-28)
    YYYYMMDD     (ADIF compact, e.g. 20260328)

--date is shorthand for --after DATE --before DATE (single-day extract).
It is mutually exclusive with --after.

Presets (--preset)
------------------
    qrz     MY_GRIDSQUARE MY_LAT MY_LON MY_STATE MY_CNTY MY_CITY MY_COUNTRY
            MY_CQ_ZONE MY_ITU_ZONE MY_DXCC MY_NAME COMMENT   [default]
    lotw    GRIDSQUARE STATE CNTY DXCC CQZ ITUZ CONT QSL_RCVD
            LOTW_QSL_RCVD APP_LOTW_2XQSL APP_LOTW_RXQSL
    n3fjp   RST_SENT RST_RCVD FREQ BAND MODE PROGRAMID LOG_PGM
    wsjtx   RST_SENT RST_RCVD FREQ BAND MODE GRIDSQUARE COMMENT

Output CSV (narrow inspection view)
------------------------------------
    field          blank - fill with ADIF field name to correct
    qso_date       YYYY-MM-DD
    time_on        HH:MM
    call           contacted station callsign
    <preset or --fields columns>
    new_value      blank - fill with corrected value

Output Excel (Full - all ADIF fields present in the source file)
-----------------------------------------------------------------
    Column order: key identity fields | preset/--fields inspection columns |
                  remaining ADIF fields alphabetically
    Inspection columns highlighted in a lighter header colour.
    Formatting: frozen header + first 4 columns, auto-filter, column widths.

Excel to ADIF (--from-xlsx)
----------------------------
    Every non-blank cell becomes an ADIF field; column header = field name.
    Columns named 'field' or 'new_value' are skipped (inspection helpers).
    QSO_DATE and TIME_ON are converted back to ADIF compact format.
    APP_QRZLOG_LOGID is preserved if present.

Requirements
------------
    pip install openpyxl requests

2026-04-14 Jim Carson (WT8P)
"""

import argparse
import csv
import logging
import re
import sys
from pathlib import Path

import qrz_common as qrz

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Preset field lists
# ---------------------------------------------------------------------------
PRESETS: dict[str, list[str]] = {
    "qrz": [
        "MY_GRIDSQUARE", "MY_LAT", "MY_LON", "MY_STATE", "MY_CNTY",
        "MY_CITY", "MY_COUNTRY", "MY_CQ_ZONE", "MY_ITU_ZONE",
        "MY_DXCC", "MY_NAME", "COMMENT",
    ],
    "lotw": [
        "GRIDSQUARE", "STATE", "CNTY", "DXCC", "CQZ", "ITUZ", "CONT",
        "QSL_RCVD", "LOTW_QSL_RCVD", "APP_LOTW_2XQSL", "APP_LOTW_RXQSL",
    ],
    "n3fjp": [
        "RST_SENT", "RST_RCVD", "FREQ", "BAND", "MODE",
        "PROGRAMID", "LOG_PGM",
    ],
    "wsjtx": [
        "RST_SENT", "RST_RCVD", "FREQ", "BAND", "MODE",
        "GRIDSQUARE", "COMMENT",
    ],
}

# Key identity columns - always first in Full Excel
KEY_FIELDS = ["QSO_DATE", "TIME_ON", "CALL", "BAND", "MODE", "FREQ"]

# Columns that are inspection helpers only - skipped on Excel->ADIF round-trip
SKIP_ON_ROUNDTRIP = {"field", "new_value", "FIELD", "NEW_VALUE"}


# ---------------------------------------------------------------------------
# Date range helpers
# ---------------------------------------------------------------------------

def _normalise_date_arg(value: str, argname: str) -> str:
    """Accept YYYY-MM-DD or YYYYMMDD; return YYYYMMDD for comparisons."""
    v = value.strip()
    if re.fullmatch(r"\d{8}", v):
        return v
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
        return v.replace("-", "")
    log.error("--%s value %r is not a recognised date (use YYYY-MM-DD or YYYYMMDD)",
              argname, value)
    sys.exit(1)


def _in_range(qso_date: str, after: str, before: str) -> bool:
    if after  and qso_date < after:  return False
    if before and qso_date > before: return False
    return True


# ---------------------------------------------------------------------------
# Column ordering for Full Excel
# ---------------------------------------------------------------------------

def _full_column_order(all_fields: set[str],
                       inspection_fields: list[str]) -> list[str]:
    """
    Build the Full Excel column order:
      1. KEY_FIELDS that are present in the data
      2. inspection_fields (preset/--fields) not already in KEY_FIELDS
      3. all remaining fields, alphabetically
    """
    used: set[str] = set()
    cols: list[str] = []

    for f in KEY_FIELDS:
        if f in all_fields:
            cols.append(f)
            used.add(f)

    for f in inspection_fields:
        if f in all_fields and f not in used:
            cols.append(f)
            used.add(f)

    for f in sorted(all_fields):
        if f not in used:
            cols.append(f)

    return cols


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------

def extract(
    records:           list[dict],
    inspection_fields: list[str],
    after:             str = "",
    before:            str = "",
) -> tuple[list[dict], list[tuple], list[str]]:
    """
    Filter records by date range.

    Returns:
        rows_narrow  -- inspection CSV rows (with field/new_value placeholders)
        filtered     -- list of (rec, adif_date, adif_time) for Excel writing
        all_columns  -- ordered full column list for Excel
    """
    all_field_names: set[str] = set()
    filtered: list[tuple] = []

    for rec in records:
        adif_date, adif_time = qrz.parse_qso_datetime(
            rec.get("QSO_DATE", ""), rec.get("TIME_ON", "")
        )
        if not _in_range(adif_date, after, before):
            continue
        filtered.append((rec, adif_date, adif_time))
        all_field_names.update(rec.keys())

    all_columns = _full_column_order(all_field_names, inspection_fields)

    rows_narrow: list[dict] = []
    for rec, adif_date, adif_time in filtered:
        hr_date, hr_time = qrz.format_qso_datetime(adif_date, adif_time)
        row: dict = {
            "field":    "",
            "qso_date": hr_date,
            "time_on":  hr_time,
            "call":     rec.get("CALL", "").upper().strip(),
            "new_value": "",
        }
        for f in inspection_fields:
            row[f] = rec.get(f, "")
        rows_narrow.append(row)

    return rows_narrow, filtered, all_columns


# ---------------------------------------------------------------------------
# CSV writer
# ---------------------------------------------------------------------------

def write_csv(rows_narrow: list[dict], inspection_fields: list[str],
              path: Path) -> None:
    cols = (["field", "qso_date", "time_on", "call"]
            + inspection_fields + ["new_value"])
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows_narrow)
    log.info("CSV  : %d rows -> %s", len(rows_narrow), path)


# ---------------------------------------------------------------------------
# Excel writer (Full)
# ---------------------------------------------------------------------------

# Approximate column widths (characters)
_WIDE_FIELDS = {
    "COMMENT", "NAME", "ADDRESS", "EMAIL", "MY_NAME", "MY_CITY",
    "COUNTRY", "MY_COUNTRY", "LOG_PGM", "PROGRAMID",
}
_NARROW_FIELDS = {
    "QSO_DATE", "TIME_ON", "TIME_OFF", "QSO_DATE_OFF",
    "BAND", "BAND_RX", "MODE", "CQZ", "ITUZ",
    "MY_CQ_ZONE", "MY_ITU_ZONE", "DXCC", "MY_DXCC", "CONT",
}


def _col_width(field: str) -> int:
    if field in _WIDE_FIELDS:    return 36
    if field in _NARROW_FIELDS:  return 12
    if field.startswith("MY_"):  return 18
    if field.startswith("APP_"): return 22
    if field in ("CALL", "STATION_CALLSIGN"): return 12
    if "LAT" in field or "LON" in field:      return 16
    return 16


def write_xlsx(
    filtered:          list[tuple],
    all_columns:       list[str],
    inspection_fields: list[str],
    path:              Path,
) -> None:
    """Write a fully-formatted Full Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl is required for --output-xlsx.  pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "QSOs"

    hdr_font      = Font(bold=True, color="FFFFFF")
    hdr_fill_dark = PatternFill("solid", fgColor="1F3864")  # dark navy
    hdr_fill_mid  = PatternFill("solid", fgColor="2E5FAC")  # medium blue (inspection cols)

    inspection_set = set(inspection_fields)

    # Header row
    for col_idx, field in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font      = hdr_font
        cell.fill      = hdr_fill_mid if field in inspection_set else hdr_fill_dark
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Data rows
    for row_idx, (rec, adif_date, adif_time) in enumerate(filtered, start=2):
        hr_date, hr_time = qrz.format_qso_datetime(adif_date, adif_time)
        for col_idx, field in enumerate(all_columns, start=1):
            if field == "QSO_DATE":
                value = hr_date
            elif field == "TIME_ON":
                value = hr_time
            elif field == "TIME_OFF":
                raw = rec.get("TIME_OFF", "")
                value = qrz.format_qso_datetime("19700101", raw)[1] if raw else ""
            elif field == "QSO_DATE_OFF":
                raw = rec.get("QSO_DATE_OFF", "")
                value = qrz.format_qso_datetime(raw, "0000")[0] if raw else ""
            else:
                value = rec.get(field, "")
            if value:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Column widths
    for col_idx, field in enumerate(all_columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(field)

    # Freeze header row + first 4 columns
    freeze_col = get_column_letter(min(5, len(all_columns) + 1))
    ws.freeze_panes = f"{freeze_col}2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    log.info("XLSX : %d rows, %d columns -> %s",
             len(filtered), len(all_columns), path)


# ---------------------------------------------------------------------------
# Excel -> ADIF round-trip
# ---------------------------------------------------------------------------

def xlsx_to_adif(xlsx_path: Path, adif_path: Path) -> None:
    """
    Convert a Full Excel workbook back to ADIF.

    Rules:
    - Every non-blank cell becomes an ADIF field (column header = field name).
    - Columns 'field' and 'new_value' are skipped (inspection helpers).
    - QSO_DATE (YYYY-MM-DD) and QSO_DATE_OFF -> YYYYMMDD.
    - TIME_ON and TIME_OFF (HH:MM) -> HHMM.
    - All other values written as-is; APP_QRZLOG_LOGID preserved.
    - Output contains only the columns present in the sheet.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl is required for --from-xlsx.  pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        log.error("Excel file appears empty: %s", xlsx_path)
        sys.exit(1)

    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    data_rows = all_rows[1:]

    records: list[dict] = []
    skipped_blank = 0

    for row in data_rows:
        rec: dict = {}
        for header, cell_val in zip(headers, row):
            if not header or header in SKIP_ON_ROUNDTRIP:
                continue
            if cell_val is None or str(cell_val).strip() == "":
                continue
            value = str(cell_val).strip()

            # Convert human-readable date/time back to ADIF compact
            if header in ("QSO_DATE", "QSO_DATE_OFF"):
                adif_d, _ = qrz.parse_qso_datetime(value)
                value = adif_d
            elif header in ("TIME_ON", "TIME_OFF"):
                # parse_qso_datetime needs a date placeholder to parse time
                _, adif_t = qrz.parse_qso_datetime("19700101", value)
                value = adif_t

            rec[header] = value

        if rec:
            records.append(rec)
        else:
            skipped_blank += 1

    if skipped_blank:
        log.info("Skipped %d blank rows", skipped_blank)

    if not records:
        log.warning("No data rows found in %s", xlsx_path)
        sys.exit(0)

    qrz.write_adif_file(records, adif_path)
    log.info("ADIF : %d records -> %s", len(records), adif_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Lightweight ADIF spreadsheet editor. "
            "Extract any ADIF file to Excel/CSV for inspection and editing, "
            "then convert the edited Excel back to ADIF (--from-xlsx)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="\n".join([
            "Presets (--preset):",
            "  qrz   : MY_GRIDSQUARE MY_LAT MY_LON MY_STATE MY_CNTY MY_CITY",
            "          MY_COUNTRY MY_CQ_ZONE MY_ITU_ZONE MY_DXCC MY_NAME COMMENT",
            "  lotw  : GRIDSQUARE STATE CNTY DXCC CQZ ITUZ CONT QSL_RCVD",
            "          LOTW_QSL_RCVD APP_LOTW_2XQSL APP_LOTW_RXQSL",
            "  n3fjp : RST_SENT RST_RCVD FREQ BAND MODE PROGRAMID LOG_PGM",
            "  wsjtx : RST_SENT RST_RCVD FREQ BAND MODE GRIDSQUARE COMMENT",
        ])
    )

    # Input mode
    mode = p.add_mutually_exclusive_group(required=True)
    mode.add_argument("--adif", metavar="FILE",
        help="Source ADIF file (QRZ, LoTW, N3FJP, WSJT-X, ...)")
    mode.add_argument("--from-xlsx", metavar="FILE",
        help="Convert an edited Full Excel workbook back to ADIF")

    # Date filtering (extract mode only)
    date_grp = p.add_mutually_exclusive_group()
    date_grp.add_argument("--date", default="", metavar="DATE",
        help="Single date shorthand for --after DATE --before DATE")
    date_grp.add_argument("--after", default="", metavar="DATE",
        help="Include QSOs on or after this date (inclusive)")
    p.add_argument("--before", default="", metavar="DATE",
        help="Include QSOs on or before this date (inclusive)")

    # Field selection (extract mode only)
    fld_grp = p.add_mutually_exclusive_group()
    fld_grp.add_argument("--preset",
        choices=list(PRESETS.keys()), default="qrz",
        help="Named inspection column preset (default: qrz)")
    fld_grp.add_argument("--fields", metavar="FIELD1,FIELD2,...",
        help="Explicit comma-separated list of inspection columns")

    # Outputs
    p.add_argument("--output-csv", default="adif_extract.csv", metavar="FILE",
        help="Narrow inspection CSV (default: adif_extract.csv)")
    p.add_argument("--no-csv", action="store_true",
        help="Suppress CSV output (e.g. when only --output-xlsx is wanted)")
    p.add_argument("--output-xlsx", default="", metavar="FILE",
        help="Full Excel output with all ADIF fields, formatted for editing")
    p.add_argument("--output-adif", default="adif_extract.adi", metavar="FILE",
        help="ADIF output for --from-xlsx round-trip (default: adif_extract.adi)")

    return p


def main() -> None:
    args = build_parser().parse_args()

    # ------------------------------------------------------------------
    # Round-trip mode: Excel -> ADIF
    # ------------------------------------------------------------------
    if args.from_xlsx:
        xlsx_path = Path(args.from_xlsx)
        if not xlsx_path.exists():
            log.error("File not found: %s", xlsx_path)
            sys.exit(1)
        log.info("=== ADIF Extract -- Excel -> ADIF ===")
        log.info("Input  : %s", xlsx_path)
        log.info("Output : %s", args.output_adif)
        xlsx_to_adif(xlsx_path, Path(args.output_adif))
        log.info("=== Done. ===")
        return

    # ------------------------------------------------------------------
    # Extract mode: ADIF -> CSV / Excel
    # ------------------------------------------------------------------
    adif_path = Path(args.adif)
    if not adif_path.exists():
        log.error("File not found: %s", adif_path)
        sys.exit(1)

    # Resolve inspection fields
    if args.fields:
        inspection_fields = [f.strip().upper()
                             for f in args.fields.split(",") if f.strip()]
    else:
        inspection_fields = PRESETS[args.preset]

    # Resolve date bounds
    if args.date:
        if args.before:
            log.error("--before cannot be used with --date")
            sys.exit(1)
        date_norm     = _normalise_date_arg(args.date, "date")
        after = before = date_norm
    else:
        after  = _normalise_date_arg(args.after,  "after")  if args.after  else ""
        before = _normalise_date_arg(args.before, "before") if args.before else ""

    if after and before and after > before:
        log.error("--after (%s) is later than --before (%s)", after, before)
        sys.exit(1)

    # Log run parameters
    log.info("=== ADIF Extract ===")
    log.info("Source  : %s", adif_path)
    preset_label = (args.preset if not args.fields
                    else f"custom ({len(inspection_fields)} fields)")
    log.info("Preset  : %s", preset_label)
    if args.date:
        log.info("Date    : %s", f"{after[:4]}-{after[4:6]}-{after[6:]}")
    else:
        if after:  log.info("After   : %s", f"{after[:4]}-{after[4:6]}-{after[6:]}")
        if before: log.info("Before  : %s", f"{before[:4]}-{before[4:6]}-{before[6:]}")

    records = qrz.parse_adif_file(adif_path)
    rows_narrow, filtered, all_columns = extract(
        records, inspection_fields, after=after, before=before
    )

    if not filtered:
        log.warning("No QSOs matched the specified criteria.")
        sys.exit(0)

    log.info("Matched : %d QSOs", len(filtered))

    if not args.no_csv:
        write_csv(rows_narrow, inspection_fields, Path(args.output_csv))

    if args.output_xlsx:
        write_xlsx(filtered, all_columns, inspection_fields, Path(args.output_xlsx))
    elif args.no_csv:
        log.warning("--no-csv specified but no --output-xlsx given; no output produced.")

    log.info("=== Done. ===")
    if not args.output_xlsx:
        log.info(
            "Tip: add --output-xlsx <file> for a full editable workbook, "
            "then --from-xlsx to convert back to ADIF after editing."
        )


if __name__ == "__main__":
    main()
